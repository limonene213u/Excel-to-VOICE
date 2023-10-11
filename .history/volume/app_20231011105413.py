# app.py
from flask import Flask, render_template, request, send_file, jsonify, redirect, url_for, send_from_directory, Blueprint, make_response
import openpyxl
import requests
import json
import wave

app = Flask(__name__, static_folder='./static', template_folder='./templates')

@app.route('/')
def index():
    return render_template('index.html')

def generate_silence(duration_ms, sample_rate=24000):
    # 無音データのサンプル数を計算
    num_samples = int(duration_ms * sample_rate / 1000)
    # すべてのサンプル値が0のバイト列を作成
    silence_data = b'\x00' * (num_samples * 2)  # 16-bit PCMなので2バイト/サンプル
    return silence_data

def process_cell(cell, speaker_id, engine_url):
    cell_value = cell.value
    if cell_value:
        # 音声化する文言と話者を指定(3で標準ずんだもんになる)
        params = (
            ('text', cell_value),
            ('speaker', speaker_id),  # ドロップダウンメニューで選択された話者IDを使用
        )
        # 音声合成用のクエリ作成
        query = requests.post(
            f'{engine_url}/audio_query',  # エンジンのURLを動的に変更
            params=params
        )
        # 音声合成を実施
        synthesis = requests.post(
            f'{engine_url}/synthesis',  # エンジンのURLを動的に変更
            headers={"Content-Type": "application/json"},
            params=params,
            data=json.dumps(query.json())
        )
        return synthesis.content  # 音声データを返す
    return b''  # ここを追加

@app.route('/upload', methods=['POST'])
def upload_file():
    file = request.files['file']
    engine_option = request.form['engine-option']
    speaker_id_a = request.form['speaker-id-a']  # A列の話者IDを取得する
    speaker_id_b = request.form['speaker-id-b']  # B列の話者IDを取得する
    single_column = request.form.get('single_column') == 'on'  # チェックボックスの値を取得する
    engine_url = 'http://host.docker.internal:50021' if engine_option == 'local' else 'http://voicevox:50021'
    
    if file:
        # Excelファイルを処理
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        wav_data = b''  # 連結された音声データを保持する
        silence_data = generate_silence(500)  # 500ミリ秒の無音データを生成する
        
        # チェックボックスがオンの場合は2行目から処理する
        for row_idx in range(2, sheet.max_row + 1):
            if single_column:  # チェックボックスがオンの場合はA列のみを処理する
                col_idx = 1
                wav_data += process_cell(sheet.cell(row=row_idx, column=col_idx), speaker_id_a, engine_url)
            else:  # チェックボックスがオフの場合はA列→B列の順に処理する
                wav_data += process_cell(sheet.cell(row=row_idx, column=1), speaker_id_a, engine_url)
                wav_data += process_cell(sheet.cell(row=row_idx, column=2), speaker_id_b, engine_url)

        # wavファイルを保存
        wav_file_path = 'output.wav'
        with wave.open(wav_file_path, 'wb') as wav_file:
            wav_file.setnchannels(1)
            wav_file.setsampwidth(2)  # 16-bit PCM
            wav_file.setframerate(24000)
            wav_file.writeframes(wav_data)

        return send_file(wav_file_path, as_attachment=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
