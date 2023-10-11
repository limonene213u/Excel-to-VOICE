# app.py
from flask import Flask, render_template, request, send_file, jsonify, redirect, url_for, send_from_directory, Blueprint, make_response
import openpyxl
import requests
import json
import wave

app = Flask(__name__, static_folder='./static', template_folder='./templates')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    file = request.files['file']
    if file:
        # Excelファイルを処理
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        wav_data = b''  # 連結された音声データを保持するにゃん
        silence = b'\x00' * 48000  # 1秒間の無音データを作成するにゃん (サンプルレート24000、2バイト/サンプルで計算)

        # A2から開始し、B列の末尾まで処理するにゃん
        for row_idx in range(2, sheet.max_row + 1):
            for col_idx in range(1, 3):  # A列とB列のインデックスは1と2にゃん
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    # 音声化する文言と話者を指定(3で標準ずんだもんになる)
                    params = (
                        ('text', cell_value),
                        ('speaker', 3),
                    )
                    # 音声合成用のクエリ作成
                    query = requests.post(
                        'http://voicevox:50021/audio_query',
                        params=params
                    )
                    # 音声合成を実施
                    synthesis = requests.post(
                        'http://voicevox:50021/synthesis',
                        headers={"Content-Type": "application/json"},
                        params=params,
                        data=json.dumps(query.json())
                    )
                    wav_data += synthesis.content  # 連結するにゃん

        # wavファイルを保存
        wav_file_path = 'output.wav'
        with wave.open(wav_file_path, 'wb') as wav_file:
            wav_file.setnchannels(1)
            wav_file.setsampwidth(2)  # 16-bit PCM
            wav_file.setframerate(24000)
            wav_file.writeframes(wav_data)

        return send_file(wav_file_path, as_attachment=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
