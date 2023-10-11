# app.py

import io
from flask import Flask, render_template, request, send_file, jsonify, send_file, redirect, url_for, send_from_directory, Blueprint, make_response
import openpyxl
import requests
import json
import wave

app = Flask(__name__, static_folder='./static', template_folder='./templates')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    file = request.files['file']
    if file:
        # Excelファイルを処理
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        wav_data = b''  # 連結された音声データを保持するにゃん
        silence = b'\x00' * 48000  # 1秒間の無音データを作成するにゃん (サンプルレート24000、2バイト/サンプルで計算)

        # A2から開始し、B列の末尾まで処理するにゃん
        for row_idx in range(2, sheet.max_row + 1):
            for col_idx in range(1, 3):  # A列とB列のインデックスは1と2にゃん
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    # 音声化する文言と話者を指定(3で標準ずんだもんになる)
                    params = (
                        ('text', cell_value),
                        ('speaker', 3),
                    )
                    # 音声合成用のクエリ作成
                    query = requests.post(
                        'http://voicevox:50021/audio_query',
                        params=params
                    )
                    # 音声合成を実施
                    synthesis = requests.post(
                        'http://voicevox:50021/synthesis',
                        headers={"Content-Type": "application/json"},
                        params=params,
                        data=json.dumps(query.json())
                    )
            wav_data += synthesis.content  # 連結するにゃん
            wav_data += silence  # 各音声データの後に無音データを追加するにゃん

        # wavファイルを保存
        wav_file_path = 'output.wav'
        with wave.open(wav_file_path, 'wb') as wav_file:
            wav_file.setnchannels(1)
            wav_file.setsampwidth(2)  # 16-bit PCM
            wav_file.setframerate(24000)
            wav_file.writeframes(wav_data)

        return send_file(wav_file_path, as_attachment=True)
    
# 既存の辞書（例）
user_dict = {}

@app.route('/add_word', methods=['POST'])
def add_word():
    word = request.form['word']
    pronunciation = request.form['pronunciation']
    accent_type = int(request.form['accent_type'])  # アクセント型も受け取るにゃ
    word_type = request.form.get('word_type', 'COMMON_NOUN')  # 単語の種類、デフォルトは普通名詞にゃ
    priority = int(request.form.get('priority', 5))  # 優先度、デフォルトは5にゃ

    user_dict[word] = pronunciation

    # VOICEVOX APIで辞書を更新する処理にゃ
    api_url = "http://voicevox:50021/user_dict_word"
    payload = {
        "surface": word,
        "pronunciation": pronunciation,
        "accent_type": accent_type,
        "word_type": word_type,
        "priority": priority
    }
    response = requests.post(api_url, json=payload)
    
    if response.status_code == 200:
        return jsonify({"message": "単語を追加しましたにゃ", "status": "success"})
    else:
        return jsonify({"message": "単語の追加に失敗したにゃ", "status": "failure"})


@app.route('/update_word', methods=['POST'])
def update_word():
    word = request.form['word']
    new_pronunciation = request.form['new_pronunciation']
    if word in user_dict:
        user_dict[word] = new_pronunciation
        # VOICEVOX APIで辞書を更新する処理（ここでは省略）
        return "単語を更新しましたにゃ"
    else:
        return "その単語は存在しないにゃ"

@app.route('/delete_word', methods=['POST'])
def delete_word():
    word = request.form['word']
    if word in user_dict:
        del user_dict[word]
        # VOICEVOX APIで辞書を更新する処理（ここでは省略）
        return "単語を削除しましたにゃ"
    else:
        return "その単語は存在しないにゃ"

@app.route('/list_words', methods=['GET'])
def list_words():
    return jsonify(user_dict)


@app.route('/sample_voice', methods=['GET'])
def sample_voice():
    sample_word = request.args.get('sample-word')

    # VOICEVOX APIで音声合成のクエリを作成
    params = {'text': sample_word, 'speaker': 3}  # 3で標準ずんだもんになる
    query_response = requests.post('http://voicevox:50021/audio_query', params=params)
    query_data = query_response.json()

    # VOICEVOX APIで音声合成を実施
    synthesis_response = requests.post(
        'http://voicevox:50021/synthesis',
        headers={"Content-Type": "application/json"},
        params=params,
        data=json.dumps(query_data)
    )
    sample_voice_data = synthesis_response.content  # バイナリデータを取得

    return send_file(
        io.BytesIO(sample_voice_data),
        mimetype='audio/wav',
        as_attachment=False
    )

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
